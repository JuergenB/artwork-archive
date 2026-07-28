#!/usr/bin/env python3
"""
Extract the column contract *and* provenance from an Artwork Archive import
template (.xlsx) into a JSON fixture the export app can assert against and
display to the curator.

AA template structure (verified against the April 2026 templates):
  Row 1  numeric column index (0..N-1), stored as formulas
  Row 2  column headers
  Row 3  helper / instruction text
  Row 4+ data

Provenance is read from the workbook's own docProps/core.xml, not the
filesystem. Filesystem mtimes only record when the file landed in our repo;
`dcterms:modified` records when Artwork Archive last revised the template,
which is the number the curator actually needs to see.

Usage:
  python3 scripts/extract-aa-template.py <template.xlsx> <artist|artwork> [out.json]

Re-run this whenever AA ships a revised template, then run the app's
aa-columns test to see exactly which headers changed.
"""
import hashlib
import json
import os
import re
import sys
import warnings
import zipfile

warnings.filterwarnings("ignore")

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required: python3 -m pip install openpyxl")

HEADER_ROW = 2
HELPER_ROW = 3

# Path prefix stripped so the recorded path stays repo-relative and portable.
REPO_MARKER = "docs/knowledge/"


def read_provenance(path: str) -> dict:
    """Pull AA's own revision metadata out of the workbook package."""
    revised_at = None
    created_at = None
    author = None

    try:
        with zipfile.ZipFile(path) as zf:
            core = zf.read("docProps/core.xml").decode("utf-8")

        def tag(name):
            match = re.search(rf"<{name}[^>]*>([^<]*)</{name}>", core)
            return match.group(1) if match else None

        revised_at = tag("dcterms:modified")
        created_at = tag("dcterms:created")
        author = tag("dc:creator")
    except (KeyError, zipfile.BadZipFile):
        pass

    with open(path, "rb") as fh:
        digest = hashlib.sha256(fh.read()).hexdigest()

    idx = path.find(REPO_MARKER)
    repo_path = path[idx:] if idx >= 0 else path

    return {
        "fileName": os.path.basename(path),
        "repoPath": repo_path,
        "revisionLabel": os.path.basename(os.path.dirname(path)),
        "revisedAt": revised_at,
        "createdAt": created_at,
        "author": author,
        "sha256": digest,
        "sizeBytes": os.path.getsize(path),
    }


def extract(path: str, entity_type: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    def row_values(idx: int) -> list:
        if ws.max_row < idx:
            return []
        return [c.value for c in ws[idx]]

    headers = row_values(HEADER_ROW)
    helpers = row_values(HELPER_ROW)

    if not headers or any(h is None for h in headers):
        sys.exit(f"{path}: row {HEADER_ROW} is not a complete header row")

    columns = []
    for i, name in enumerate(headers):
        helper = helpers[i] if i < len(helpers) else None
        columns.append(
            {
                "index": i,
                "name": str(name),
                "helperText": "" if helper is None else str(helper),
            }
        )

    return {
        "entityType": entity_type,
        "provenance": read_provenance(path),
        "sheetName": ws.title,
        "headerRow": HEADER_ROW,
        "helperRow": HELPER_ROW,
        "firstDataRow": HELPER_ROW + 1,
        "columnCount": len(columns),
        "columns": columns,
    }


def main() -> None:
    if len(sys.argv) < 3:
        sys.exit(__doc__)

    template_path, entity_type = sys.argv[1], sys.argv[2]
    if entity_type not in ("artist", "artwork"):
        sys.exit("entityType must be 'artist' or 'artwork'")

    result = extract(template_path, entity_type)
    payload = json.dumps(result, indent=2, ensure_ascii=False) + "\n"

    if len(sys.argv) >= 4:
        with open(sys.argv[3], "w", encoding="utf-8") as fh:
            fh.write(payload)
        prov = result["provenance"]
        print(
            f"{result['columnCount']} columns  "
            f"revised={prov['revisedAt']}  sha256={prov['sha256'][:12]}…  → {sys.argv[3]}"
        )
    else:
        sys.stdout.write(payload)


if __name__ == "__main__":
    main()
